#!/usr/bin/env python3
"""
Build Alleato Group billing workbooks (Form APP / CS / CO).

Variants:
  owner : Contractor -> Owner        (prime contract application for payment)
  sub   : Subcontractor -> Contractor (subcontract / commitment pay app)

The Schedule of Values sheet keeps the industry continuation-sheet column letters
(A,B,C,D,E,F,G,G/C,H,I) but adds a Budget Code column and explodes column I into
the 5a (completed work) / 5b (stored materials) retainage split so Line 5 of the
Payment Application is auditable.

All cross-sheet references use workbook defined names, so row shifts on Setup
cannot silently break the formulas on the other sheets.
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

MONEY = '#,##0.00;[Red](#,##0.00)'
PCT = '0.0%'
PCT100 = '0.0"%"'   # platform stores percentages out of 100, not as fractions
DATE = 'mm/dd/yyyy'

# ---- Alleato Group brand palette
ORANGE = 'FD5602'      # Primary Orange
ORANGE2 = 'DB802E'     # Secondary Orange (tagline)
BLACK = '000000'
DKGREY = '454545'      # logo text / header banner
LTGREY = '9C9998'

HDR_FILL = PatternFill('solid', fgColor=DKGREY)      # column header banners
SUB_FILL = PatternFill('solid', fgColor='EDEBEA')    # light grey section bands
TOT_FILL = PatternFill('solid', fgColor='FDEFE7')    # 8% Primary Orange — totals
ACCENT_FILL = PatternFill('solid', fgColor=ORANGE)   # orange accent rule
IN_FILL = PatternFill('solid', fgColor='FFFFFF')
LOCK_FILL = PatternFill('solid', fgColor='F7F7F6')

THIN = Side(style='thin', color='C9C6C5')
MED = Side(style='medium', color=DKGREY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BRAND_TAGLINE = 'Your partner from the ground up.'
BRAND_WEB = 'www.alleatogroup.com'
BRAND_HQ = '8383 Craig Street, Suite 150, Indianapolis, IN 46250'
BRAND_FL = '701 94th Avenue North, Suite 118, St. Petersburg, FL 33702'

import os
_HERE = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(_HERE, 'brand', 'Alleato-Logo-Dark.png')
_LOGO_FOR_ASPECT = LOGO_PATH


def _logo_aspect(path=None):
    """Width / height of the logo asset, read from the file so resizing it is safe."""
    try:
        from PIL import Image
        with Image.open(path or _LOGO_FOR_ASPECT) as im:
            return im.width / im.height
    except Exception:
        return 1956 / 593.0


F_TITLE = Font(name='Calibri', size=16, bold=True, color=BLACK)
F_SUB = Font(name='Calibri', size=10, italic=True, color=DKGREY)
F_HDR = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
F_LBL = Font(name='Calibri', size=10, bold=True)
F_BODY = Font(name='Calibri', size=10)
F_SMALL = Font(name='Calibri', size=8, color=DKGREY)
F_TOT = Font(name='Calibri', size=10, bold=True)
F_BIG = Font(name='Calibri', size=12, bold=True)
F_SECT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

# ---- sheet names (no third-party form references anywhere in the output) ----
SH_SOV = 'Schedule of Values'      # working data-entry sheet
SH_CS = 'Continuation Sheet'       # strict transmittal sheet — Form CS
SH_APP = 'Payment Application'     # Form APP
SH_CO = 'Change Order'             # Form CO

# --------------------------------------------------------------- SOV layout
SOV_COLS = [
    ('A', 'A',   'ITEM\nNO.',                                    7),
    ('B', 'B',   'DESCRIPTION OF WORK',                          38),
    ('C', '',    'BUDGET CODE',                                  14),
    ('D', 'C',   'SCHEDULED\nVALUE',                             14),
    ('E', 'D',   'WORK COMPLETED\nFROM PREVIOUS\nAPPLICATION',   15),
    ('F', 'E',   'WORK COMPLETED\nTHIS PERIOD',                  15),
    ('G', 'F',   'MATERIALS\nPRESENTLY\nSTORED',                 14),
    ('H', 'G',   'TOTAL COMPLETED\nAND STORED\nTO DATE',         16),
    ('I', 'G/C', '%',                                             7),
    ('J', 'H',   'BALANCE TO\nFINISH',                           14),
    ('K', 'I',   'RET %\nWORK',                                   8),
    ('L', 'I',   'RET $ WORK\nPRIOR',                            13),
    ('M', 'I',   'RET $ WORK\nTHIS PERIOD',                      13),
    ('N', 'I',   'RET $ WORK\nRELEASED',                         13),
    ('O', 'I',   'RET %\nMATL',                                   8),
    ('P', 'I',   'RET $ MATL\nPRIOR',                            13),
    ('Q', 'I',   'RET $ MATL\nTHIS PERIOD',                      13),
    ('R', 'I',   'RET $ MATL\nRELEASED',                         13),
    ('S', 'I',   'TOTAL RETAINAGE\nHELD',                        15),
]
NCOL = len(SOV_COLS)

BASE_FIRST, BASE_LAST = 12, 46
BASE_SUB = 47
CO_SECT = 49
CO_FIRST, CO_LAST = 50, 64
CO_SUB = 65
GRAND = 67

CO_LOG_FIRST, CO_LOG_LAST = 7, 46


def variant_words(variant):
    if variant == 'owner':
        return dict(
            payee='CONTRACTOR', payer='OWNER',
            title='APPLICATION AND CERTIFICATE FOR PAYMENT',
            subtitle='Form APP  ·  Prime Contract  ·  Contractor to Owner',
            to_label='TO OWNER', from_label='FROM CONTRACTOR', via_label='VIA ARCHITECT',
            contract_label='PRIME CONTRACT',
            third_party='ARCHITECT / DESIGN PROFESSIONAL', third_role='Architect',
            certifier='ARCHITECT', cert_title='ARCHITECT’S CERTIFICATE FOR PAYMENT',
            co_title='Owner ↔ Contractor',
            co_parties=['OWNER', 'ARCHITECT', 'CONTRACTOR'],
        )
    return dict(
        payee='SUBCONTRACTOR', payer='CONTRACTOR',
        title='SUBCONTRACTOR APPLICATION FOR PAYMENT',
        subtitle='Form APP  ·  Subcontract / Commitment  ·  Subcontractor to Contractor',
        to_label='TO CONTRACTOR', from_label='FROM SUBCONTRACTOR', via_label='VIA PROJECT MANAGER',
        contract_label='SUBCONTRACT',
        third_party='PROJECT MANAGEMENT', third_role='Project Manager',
        certifier='CONTRACTOR', cert_title='CONTRACTOR’S APPROVAL FOR PAYMENT',
        co_title='Contractor ↔ Subcontractor',
        co_parties=['CONTRACTOR', 'SUBCONTRACTOR'],
    )


def style_input(cell, fmt=None):
    cell.fill = IN_FILL
    cell.border = BOX
    cell.font = F_BODY
    if fmt:
        cell.number_format = fmt


def style_calc(cell, fmt=MONEY):
    cell.fill = LOCK_FILL
    cell.border = BOX
    cell.font = F_BODY
    cell.number_format = fmt


def brand_mast(ws, last_col, accent_row=3, logo_col=None, logo_h=42):
    """Float the Alleato logo top-right, add the tagline, and lay an orange rule."""
    if os.path.isfile(LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(LOGO_PATH)
            img.height = logo_h
            img.width = int(logo_h * _logo_aspect())
            anchor_col = logo_col or max(1, last_col - 3)
            img.anchor = f'{get_column_letter(anchor_col)}1'
            ws.add_image(img)
        except Exception:
            pass
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    for col in range(1, last_col + 1):
        ws.cell(accent_row, col).fill = ACCENT_FILL
    ws.row_dimensions[accent_row].height = 3.5


def brand_print(ws, doc=None):
    """Excel page header/footer so every printed page carries the brand."""
    ws.oddFooter.left.text = BRAND_TAGLINE
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.font = 'Calibri,Italic'
    ws.oddFooter.left.color = ORANGE2
    ws.oddFooter.center.text = 'Page &P of &N'
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.color = DKGREY
    ws.oddFooter.right.text = BRAND_WEB
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.font = 'Calibri,Bold'
    ws.oddFooter.right.color = DKGREY
    if doc:
        ws.oddHeader.right.text = f'Alleato Group  ·  Form {doc}'
        ws.oddHeader.right.size = 8
        ws.oddHeader.right.color = LTGREY


def brand_foot(ws, row, last_col):
    t = ws.cell(row, 1, BRAND_TAGLINE)
    t.font = Font(name='Calibri', size=9, italic=True, color=ORANGE2)
    w = ws.cell(row, max(1, last_col - 2), BRAND_WEB)
    w.font = Font(name='Calibri', size=9, bold=True, color=DKGREY)
    w.alignment = Alignment(horizontal='right')
    for col in range(1, last_col + 1):
        ws.cell(row + 1, col).fill = ACCENT_FILL
    ws.row_dimensions[row + 1].height = 3.5


def banner(ws, row, col_first, col_last, text):
    c = ws.cell(row, col_first, text)
    c.font = F_SECT
    for col in range(col_first, col_last + 1):
        ws.cell(row, col).fill = HDR_FILL
    if col_last > col_first:
        ws.merge_cells(start_row=row, start_column=col_first, end_row=row, end_column=col_last)


# ------------------------------------------------------------------- 1. SETUP
def build_setup(wb, v, variant):
    ws = wb.create_sheet('Setup')
    ws.sheet_view.showGridLines = False
    for k, w in {'A': 3, 'B': 34, 'C': 44, 'D': 3, 'E': 62}.items():
        ws.column_dimensions[k].width = w

    brand_mast(ws, 5, accent_row=1, logo_col=5, logo_h=40)
    brand_print(ws)
    ws['B2'] = 'PROJECT & CONTRACT SETUP'
    ws['B2'].font = F_TITLE
    ws['B3'] = 'Fill the white cells once per contract. Every other sheet reads from here.'
    ws['B3'].font = F_SUB

    rows = [
        ('SECTION', 'PROJECT', None, None),
        ('Project Name', '', 'PrjName', None),
        ('Project Number', '', 'PrjNumber', None),
        ('Project Address', '', 'PrjAddress', None),
        ('Alleato Project ID (projects.id)', '', 'PrjId', None),
        ('SECTION', v['payer'].title() + ' — Paying Party', None, None),
        (v['payer'].title() + ' Name', '', 'PayerName', None),
        (v['payer'].title() + ' Address', '', 'PayerAddress', None),
        ('SECTION', v['payee'].title() + ' — Paid Party', None, None),
        (v['payee'].title() + ' Name', '', 'PayeeName', None),
        (v['payee'].title() + ' Address', '', 'PayeeAddress', None),
        ('Vendor ID (companies.id)', '', 'VendorId', None),
        ('SECTION', v['third_party'], None, None),
        (v['third_role'] + ' Name', '', 'ThirdName', None),
        ('SECTION', 'CONTRACT', None, None),
        (v['contract_label'].title() + ' Number', '', 'ContractNo', None),
        (v['contract_label'].title() + ' Date', '', 'ContractDate', DATE),
        ('Original Contract Sum', 0, 'OrigContractSum', MONEY),
        ('Alleato Contract ID', '', 'ContractId', None),
        ('SECTION', 'RETAINAGE — drives Lines 5a / 5b', None, None),
        ('Retainage % — Completed Work (5a)', 10.0, 'RetPctWork', PCT100),
        ('Retainage % — Stored Materials (5b)', 10.0, 'RetPctMatl', PCT100),
        ('SECTION', 'CURRENT APPLICATION', None, None),
        ('Application No.', 1, 'AppNo', '#,##0'),
        ('Application Date', '', 'AppDate', DATE),
        ('Period From', '', 'PeriodFrom', DATE),
        ('Period To', '', 'PeriodTo', DATE),
    ]

    r = 5
    for label, val, key, fmt in rows:
        if label == 'SECTION':
            ws.cell(r, 2, val.upper()).font = F_HDR
            for col in (2, 3):
                ws.cell(r, col).fill = HDR_FILL
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            r += 1
            continue
        lc = ws.cell(r, 2, label)
        lc.font = F_LBL
        lc.border = BOX
        lc.alignment = Alignment(vertical='center')
        tgt = ws.cell(r, 3, val if val != '' else None)
        style_input(tgt, fmt)
        wb.defined_names[key] = DefinedName(key, attr_text=f'Setup!$C${r}')
        r += 1

    guide = [
        ('h', 'HOW THIS WORKBOOK IS WIRED'),
        ('p', '1.  Setup — project, parties, contract sum, retainage rates. Type here only.'),
        ('p', '2.  CO Log — log every change order. Feeds the change order summary.'),
        ('p', '3.  Schedule of Values — the SOV and the monthly data entry.'),
        ('p', '4.  Payment Application — reads from the SOV + CO Log. Only Line 7 is typed.'),
        ('p', '5.  Change Order — one per change. Copy the sheet as needed.'),
        ('h', 'MONTHLY ROLLOVER'),
        ('p', '• Copy the workbook, bump Application No., set the new period dates.'),
        ('p', '• On the SOV, paste last month’s column H into column E as VALUES, clear F and G.'),
        ('p', '• On the Payment Application, set Line 7 to last month’s Line 6.'),
        ('p', '• On CO Log, flip "In Prior Application" to Y for change orders already billed.'),
        ('h', 'RETAINAGE'),
        ('p', '• Rates default from here but override per line in SOV columns K and O —'),
        ('p', '  that is how you handle a reduced rate or release on a single SOV line.'),
        ('p', '• Never adjust Line 5 directly; it must always foot to the SOV.'),
        ('h', 'ALLEATO GROUP standard billing forms — see the README sheet.'),
    ]
    rr = 5
    for kind, text in guide:
        c = ws.cell(rr, 5, text)
        if kind == 'h':
            c.font = F_LBL
        else:
            c.font = F_SMALL
        c.alignment = Alignment(vertical='top')
        rr += 1


# ------------------------------------------------------------------ 2. CO LOG
def build_co_log(wb, v):
    ws = wb.create_sheet('CO Log')
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([10, 40, 14, 14, 16, 16, 14, 22, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    brand_mast(ws, 9, accent_row=4, logo_col=7, logo_h=40)
    brand_print(ws)
    ws['A1'] = 'CHANGE ORDER LOG'
    ws['A1'].font = F_TITLE
    ws['A2'] = ('One row per change order. Only rows with Status = Approved or Executed are picked up. '
                '"In Prior Application" = Y once the change order has appeared on a certified application.')
    ws['A2'].font = F_SUB

    heads = ['CO NO.', 'DESCRIPTION', 'DATE\nAPPROVED', 'STATUS', 'ADDITIONS ($)',
             'DEDUCTIONS ($)', 'SCHEDULE\nIMPACT (DAYS)', 'IN PRIOR\nAPPLICATION? (Y/N)', 'NET ($)']
    for i, h in enumerate(heads, start=1):
        c = ws.cell(6, i, h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[6].height = 32

    dv_status = DataValidation(type='list', formula1='"Approved,Executed,Pending,Draft,Void"', allow_blank=True)
    dv_yn = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_yn)

    for r in range(CO_LOG_FIRST, CO_LOG_LAST + 1):
        style_input(ws.cell(r, 1))
        style_input(ws.cell(r, 2))
        style_input(ws.cell(r, 3), DATE)
        style_input(ws.cell(r, 4))
        style_input(ws.cell(r, 5), MONEY)
        style_input(ws.cell(r, 6), MONEY)
        style_input(ws.cell(r, 7), '#,##0')
        style_input(ws.cell(r, 8))
        style_calc(ws.cell(r, 9, f'=IF(COUNT(E{r}:F{r})=0,"",N(E{r})-N(F{r}))'))
        dv_status.add(ws.cell(r, 4))
        dv_yn.add(ws.cell(r, 8))

    tr = CO_LOG_LAST + 1
    ws.cell(tr, 2, 'TOTALS — APPROVED / EXECUTED ONLY').font = F_TOT
    for col in range(1, 10):
        ws.cell(tr, col).fill = TOT_FILL
        ws.cell(tr, col).border = Border(top=MED, bottom=MED, left=THIN, right=THIN)
    for col in (5, 6, 9):
        L = get_column_letter(col)
        c = ws.cell(tr, col,
                    f'=SUMPRODUCT((($D${CO_LOG_FIRST}:$D${CO_LOG_LAST}="Approved")'
                    f'+($D${CO_LOG_FIRST}:$D${CO_LOG_LAST}="Executed"))'
                    f'*N(${L}${CO_LOG_FIRST}:${L}${CO_LOG_LAST}))')
        c.number_format = MONEY
        c.font = F_TOT

    ws.cell(tr + 2, 2, 'Approved in previous applications  (In Prior = Y)').font = F_LBL
    ws.cell(tr + 3, 2, 'Approved this application  (In Prior = N)').font = F_LBL
    for i, flag in enumerate(('Y', 'N')):
        rr = tr + 2 + i
        for col in (5, 6):
            L = get_column_letter(col)
            c = ws.cell(rr, col,
                        f'=SUMIFS(${L}${CO_LOG_FIRST}:${L}${CO_LOG_LAST},'
                        f'$H${CO_LOG_FIRST}:$H${CO_LOG_LAST},"{flag}",'
                        f'$D${CO_LOG_FIRST}:$D${CO_LOG_LAST},"Approved")'
                        f'+SUMIFS(${L}${CO_LOG_FIRST}:${L}${CO_LOG_LAST},'
                        f'$H${CO_LOG_FIRST}:$H${CO_LOG_LAST},"{flag}",'
                        f'$D${CO_LOG_FIRST}:$D${CO_LOG_LAST},"Executed")')
            style_calc(c)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '6:6'

    wb.defined_names['CO_NetTotal'] = DefinedName('CO_NetTotal', attr_text=f"'CO Log'!$I${tr}")
    wb.defined_names['CO_PrevAdd'] = DefinedName('CO_PrevAdd', attr_text=f"'CO Log'!$E${tr + 2}")
    wb.defined_names['CO_PrevDed'] = DefinedName('CO_PrevDed', attr_text=f"'CO Log'!$F${tr + 2}")
    wb.defined_names['CO_ThisAdd'] = DefinedName('CO_ThisAdd', attr_text=f"'CO Log'!$E${tr + 3}")
    wb.defined_names['CO_ThisDed'] = DefinedName('CO_ThisDed', attr_text=f"'CO Log'!$F${tr + 3}")


# ------------------------------------------------------- 3. Schedule of Values
def build_schedule_of_values(wb, v, variant):
    ws = wb.create_sheet(SH_SOV)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'C12'
    for col, letter, head, w in SOV_COLS:
        ws.column_dimensions[col].width = w

    brand_mast(ws, NCOL, accent_row=3, logo_col=NCOL - 3, logo_h=40)
    brand_print(ws, 'CS')
    ws['A1'] = 'CONTINUATION SHEET — SCHEDULE OF VALUES'
    ws['A1'].font = F_TITLE
    ws['A2'] = 'Working schedule of values. Feeds the Continuation Sheet and the Payment Application.'
    ws['A2'].font = F_SUB

    for lc, lbl, vc, f, fmt in [
        ('B4', 'PROJECT:', 'C4', '=IF(PrjName="","",PrjName)', None),
        ('B5', v['payee'] + ':', 'C5', '=IF(PayeeName="","",PayeeName)', None),
        ('B6', v['contract_label'] + ' NO.:', 'C6', '=IF(ContractNo="","",ContractNo)', None),
        ('G4', 'APPLICATION NO.:', 'H4', '=AppNo', '#,##0'),
        ('G5', 'PERIOD TO:', 'H5', '=IF(PeriodTo="","",PeriodTo)', DATE),
        ('G6', 'APP. DATE:', 'H6', '=IF(AppDate="","",AppDate)', DATE),
    ]:
        ws[lc] = lbl
        ws[lc].font = F_LBL
        ws[vc] = f
        ws[vc].font = F_BODY
        if fmt:
            ws[vc].number_format = fmt

    for col, letter, head, w in SOV_COLS:
        c = ws[f'{col}9']
        c.value = letter
        c.font = Font(name='Calibri', size=8, bold=True, color=ORANGE)
        c.alignment = Alignment(horizontal='center')
        c.fill = SUB_FILL
        c.border = BOX
        h = ws[f'{col}10']
        h.value = head
        h.font = F_HDR
        h.fill = HDR_FILL
        h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        h.border = BOX
    ws.row_dimensions[9].height = 13
    ws.row_dimensions[10].height = 46

    def section(row, label):
        ws.cell(row, 1, label).font = Font(name='Calibri', size=9, bold=True, color=DKGREY)
        for i in range(1, NCOL + 1):
            ws.cell(row, i).fill = SUB_FILL
            ws.cell(row, i).border = BOX
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    section(11, 'BASE CONTRACT')
    section(CO_SECT, 'APPROVED CHANGE ORDERS')

    def data_rows(first, last):
        for r in range(first, last + 1):
            blank = f'COUNTA(A{r}:C{r})+COUNT(D{r}:G{r})=0'
            for c in (1, 2, 3):
                style_input(ws.cell(r, c))
            for c in (4, 5, 6, 7):
                style_input(ws.cell(r, c), MONEY)
            style_calc(ws.cell(r, 8, f'=IF({blank},"",N(E{r})+N(F{r})+N(G{r}))'))
            style_calc(ws.cell(r, 9, f'=IF(N(D{r})=0,"",H{r}/D{r}*100)'), PCT100)
            style_calc(ws.cell(r, 10, f'=IF({blank},"",N(D{r})-N(H{r}))'))
            style_input(ws.cell(r, 11, f'=IF({blank},"",RetPctWork)'), PCT100)
            style_input(ws.cell(r, 12), MONEY)
            style_calc(ws.cell(r, 13, f'=IF({blank},"",N(F{r})*N(K{r})/100)'))
            style_input(ws.cell(r, 14), MONEY)
            style_input(ws.cell(r, 15, f'=IF({blank},"",RetPctMatl)'), PCT100)
            style_input(ws.cell(r, 16), MONEY)
            style_calc(ws.cell(r, 17, f'=IF({blank},"",N(G{r})*N(O{r})/100)'))
            style_input(ws.cell(r, 18), MONEY)
            style_calc(ws.cell(r, 19,
                               f'=IF({blank},"",N(L{r})+N(M{r})-N(N{r})+N(P{r})+N(Q{r})-N(R{r}))'))

    data_rows(BASE_FIRST, BASE_LAST)
    data_rows(CO_FIRST, CO_LAST)

    SUM_COLS = (4, 5, 6, 7, 8, 10, 12, 13, 14, 16, 17, 18, 19)

    def subtotal(row, label, first, last):
        ws.cell(row, 2, label).font = F_TOT
        for i in range(1, NCOL + 1):
            cc = ws.cell(row, i)
            cc.fill = SUB_FILL
            cc.border = Border(top=MED, bottom=THIN, left=THIN, right=THIN)
        for col in SUM_COLS:
            L = get_column_letter(col)
            t = ws.cell(row, col, f'=SUM({L}{first}:{L}{last})')
            t.number_format = MONEY
            t.font = F_TOT
        p = ws.cell(row, 9, f'=IF(N(D{row})=0,"",H{row}/D{row}*100)')
        p.number_format = PCT100
        p.font = F_TOT

    subtotal(BASE_SUB, 'SUBTOTAL — BASE CONTRACT', BASE_FIRST, BASE_LAST)
    subtotal(CO_SUB, 'SUBTOTAL — APPROVED CHANGE ORDERS', CO_FIRST, CO_LAST)

    r = GRAND
    ws.cell(r, 2, 'GRAND TOTAL — CARRIES TO THE PAYMENT APPLICATION').font = Font(name='Calibri', size=11, bold=True, color=ORANGE)
    for i in range(1, NCOL + 1):
        cc = ws.cell(r, i)
        cc.fill = TOT_FILL
        cc.border = Border(top=MED, bottom=MED, left=THIN, right=THIN)
    for col in SUM_COLS:
        L = get_column_letter(col)
        t = ws.cell(r, col, f'=N({L}{BASE_SUB})+N({L}{CO_SUB})')
        t.number_format = MONEY
        t.font = Font(name='Calibri', size=11, bold=True)
    p = ws.cell(r, 9, f'=IF(N(D{r})=0,"",H{r}/D{r}*100)')
    p.number_format = PCT100
    p.font = Font(name='Calibri', size=11, bold=True)

    for nm, col in [('SOV_ScheduledValue', 'D'), ('SOV_WorkPrevious', 'E'),
                    ('SOV_WorkThisPeriod', 'F'), ('SOV_MaterialsStored', 'G'),
                    ('SOV_TotalCompleted', 'H'), ('SOV_BalanceToFinish', 'J'),
                    ('SOV_RetWorkPrior', 'L'), ('SOV_RetWorkThis', 'M'),
                    ('SOV_RetWorkReleased', 'N'), ('SOV_RetMatlPrior', 'P'),
                    ('SOV_RetMatlThis', 'Q'), ('SOV_RetMatlReleased', 'R'),
                    ('SOV_RetTotal', 'S')]:
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"'{SH_SOV}'!${col}${GRAND}")
    wb.defined_names['SOV_ChangeOrders'] = DefinedName('SOV_ChangeOrders', attr_text=f"'{SH_SOV}'!$D${CO_SUB}")

    ws.cell(r + 2, 2, 'TIE-OUT — Scheduled Value total vs Line 3, Contract Sum to Date:').font = F_LBL
    chk = ws.cell(r + 2, 8,
                  '=IF(ROUND(SOV_ScheduledValue-APP_ContractSumToDate,2)=0,'
                  '"OK — ties","OUT OF BALANCE by "&TEXT(SOV_ScheduledValue-APP_ContractSumToDate,"#,##0.00"))')
    chk.font = F_TOT
    ws.merge_cells(start_row=r + 2, start_column=8, end_row=r + 2, end_column=14)

    ws.print_title_rows = '9:10'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ------------------------------------------- 3b. Continuation Sheet (Form CS)
# Columns A-I only, the standard continuation-sheet layout. Every cell is a
# formula pointing at the Schedule of Values sheet — nothing is typed here.
# This is the handout / transmittal copy.
CS_COLS = [
    ('A', 'A', 'ITEM\nNO.',                                     8,  'A',  None),
    ('B', 'B', 'DESCRIPTION OF WORK',                           50,  'B',  None),
    ('C', 'C', 'SCHEDULED\nVALUE',                              16,  'D',  MONEY),
    ('D', 'D', 'FROM PREVIOUS\nAPPLICATION\n(D + E)',            17,  'E',  MONEY),
    ('E', 'E', 'THIS PERIOD',                                   16,  'F',  MONEY),
    ('F', 'F', 'MATERIALS\nPRESENTLY STORED\n(NOT IN D OR E)',   18,  'G',  MONEY),
    ('G', 'G', 'TOTAL COMPLETED\nAND STORED\nTO DATE (D+E+F)',   18,  'H',  MONEY),
    ('H', '',  '%\n(G ÷ C)',                                     9,  'I',  PCT100),
    ('I', 'H', 'BALANCE TO\nFINISH (C – G)',                    16,  'J',  MONEY),
    ('J', 'I', 'RETAINAGE\n(IF VARIABLE RATE)',                 17,  'S',  MONEY),
]
WC_FIRST, WC_LAST = 4, 5      # columns D and E sit under the WORK COMPLETED band
G_SPAN_START = 6              # 0-based index of column G; its letter spans G and the % column


def build_continuation_sheet(wb, v, variant):
    ws = wb.create_sheet(SH_CS)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'C12'
    NA = len(CS_COLS)
    for col, letter, head, w, srcw, fmt in CS_COLS:
        ws.column_dimensions[col].width = w

    brand_mast(ws, NA, accent_row=3, logo_col=NA - 2, logo_h=40)
    brand_print(ws, 'CS')
    ws['A1'] = 'CONTINUATION SHEET'
    ws['A1'].font = F_TITLE
    ws['A2'] = ('Columns A through I only. Every value is a formula from the Schedule of Values '
                'sheet — do not type here. This is the transmittal copy.')
    ws['A2'].font = F_SUB

    note = ws.cell(4, 1,
                   'The Application and Certificate for Payment containing the '
                   + v['payee'].title() + '\u2019s signed certification is attached. '
                   'Use Column I on Contracts where variable retainage for line items may apply.')
    note.font = F_SMALL
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=4)

    third_lbl = "ARCHITECT’S PROJECT NO.:" if variant == 'owner' else 'PROJECT NO.:'
    for i, (lbl, f, fmt) in enumerate([
        ('APPLICATION NO.:', '=IF(AppNo="","",AppNo)', '#,##0'),
        ('APPLICATION DATE:', '=IF(AppDate="","",AppDate)', DATE),
        ('PERIOD TO:', '=IF(PeriodTo="","",PeriodTo)', DATE),
        (third_lbl, '=IF(PrjNumber="","",PrjNumber)', None),
    ]):
        r = 4 + i
        lc = ws.cell(r, 6, lbl)
        lc.font = F_LBL
        lc.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        vc = ws.cell(r, 8, f)
        vc.font = F_BODY
        vc.border = Border(bottom=THIN)
        if fmt:
            vc.number_format = fmt
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)

    # row 8 = column letters (G spans the total column and the unlettered % column)
    i = 0
    while i < NA:
        letter = CS_COLS[i][1]
        span = 2 if i == G_SPAN_START else 1
        for k in range(span):
            c = ws.cell(8, i + 1 + k)
            c.fill = SUB_FILL
            c.border = BOX
        c = ws.cell(8, i + 1, letter)
        c.font = Font(name='Calibri', size=8, bold=True, color=ORANGE)
        c.alignment = Alignment(horizontal='center')
        if span == 2:
            ws.merge_cells(start_row=8, start_column=i + 1, end_row=8, end_column=i + 2)
        i += span
    ws.row_dimensions[8].height = 13

    # row 9 = WORK COMPLETED band over columns D and E only
    for col in range(1, NA + 1):
        cc = ws.cell(9, col)
        cc.border = BOX
        if col in (WC_FIRST, WC_LAST):
            cc.fill = HDR_FILL
        else:
            cc.fill = HDR_FILL
    wc = ws.cell(9, WC_FIRST, 'WORK COMPLETED')
    wc.font = F_HDR
    wc.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=9, start_column=WC_FIRST, end_row=9, end_column=WC_LAST)
    for col in range(1, NA + 1):
        if col in (WC_FIRST, WC_LAST):
            continue
        ws.merge_cells(start_row=9, start_column=col, end_row=10, end_column=col)
    ws.row_dimensions[9].height = 15

    # row 10 = column headings
    for i, (col, letter, head, w, srcw, fmt) in enumerate(CS_COLS, start=1):
        target = ws.cell(9, i) if i not in (WC_FIRST, WC_LAST) else ws.cell(10, i)
        target.value = head
        target.font = F_HDR
        target.fill = HDR_FILL
        target.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        target.border = BOX
        if i in (WC_FIRST, WC_LAST):
            ws.cell(10, i).border = BOX
    ws.row_dimensions[10].height = 42

    def mirror(row, fill=None, bold=False, border=None):
        for i, (col, letter, head, w, srcw, fmt) in enumerate(CS_COLS, start=1):
            c = ws.cell(row, i, f'=IF(\'{SH_SOV}\'!{srcw}{row}="","",\'{SH_SOV}\'!{srcw}{row})')
            c.border = border or BOX
            c.font = F_TOT if bold else F_BODY
            if fmt:
                c.number_format = fmt
            if fill:
                c.fill = fill

    def label_row(row, label, fill):
        for i in range(1, NA + 1):
            ws.cell(row, i).fill = fill
            ws.cell(row, i).border = BOX
        c = ws.cell(row, 1, label)
        c.font = Font(name='Calibri', size=9, bold=True, color=DKGREY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    label_row(11, 'BASE CONTRACT', SUB_FILL)
    for r in range(BASE_FIRST, BASE_LAST + 1):
        mirror(r)
    mirror(BASE_SUB, fill=SUB_FILL, bold=True,
           border=Border(top=MED, bottom=THIN, left=THIN, right=THIN))
    ws.cell(BASE_SUB, 2, 'SUBTOTAL — BASE CONTRACT').font = F_TOT

    label_row(CO_SECT, 'APPROVED CHANGE ORDERS', SUB_FILL)
    for r in range(CO_FIRST, CO_LAST + 1):
        mirror(r)
    mirror(CO_SUB, fill=SUB_FILL, bold=True,
           border=Border(top=MED, bottom=THIN, left=THIN, right=THIN))
    ws.cell(CO_SUB, 2, 'SUBTOTAL — APPROVED CHANGE ORDERS').font = F_TOT

    mirror(GRAND, fill=TOT_FILL, bold=True,
           border=Border(top=MED, bottom=MED, left=THIN, right=THIN))
    g = ws.cell(GRAND, 2, 'GRAND TOTAL')
    g.font = Font(name='Calibri', size=11, bold=True, color=ORANGE)
    for i in range(1, NA + 1):
        ws.cell(GRAND, i).font = Font(name='Calibri', size=11, bold=True)
    g.font = Font(name='Calibri', size=11, bold=True, color=ORANGE)

    n = ws.cell(GRAND + 2, 1,
                'The RETAINAGE column shows total retainage held (completed work + stored materials, net of '
                'releases) — column I, "Retainage (if variable rate)". The 5a / 5b breakdown behind '
                'it is on the Schedule of Values sheet, columns K through S.')
    n.font = F_SMALL
    n.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=GRAND + 2, start_column=1, end_row=GRAND + 3, end_column=NA)

    ws.print_title_rows = '9:10'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ----------------------------------------------------- 4. Payment Application
def build_payment_application(wb, v, variant):
    ws = wb.create_sheet(SH_APP)
    ws.sheet_view.showGridLines = False
    for k, w in {'A': 3, 'B': 6, 'C': 46, 'D': 11, 'E': 18, 'F': 4, 'G': 24, 'H': 17, 'I': 17}.items():
        ws.column_dimensions[k].width = w

    brand_mast(ws, 9, accent_row=3, logo_col=7, logo_h=40)
    brand_print(ws, 'APP')
    ws['B1'] = v['title']
    ws['B1'].font = F_TITLE
    ws['B2'] = v['subtitle']
    ws['B2'].font = F_SUB

    party = [
        (4,  v['to_label'] + ':',            'PayerName',    'APPLICATION NO.:',             'AppNo',        '#,##0'),
        (5,  '',                             'PayerAddress', 'APPLICATION DATE:',            'AppDate',      DATE),
        (6,  v['from_label'] + ':',          'PayeeName',    'PERIOD FROM:',                 'PeriodFrom',   DATE),
        (7,  '',                             'PayeeAddress', 'PERIOD TO:',                   'PeriodTo',     DATE),
        (8,  'PROJECT:',                     'PrjName',      v['contract_label'] + ' NO.:',  'ContractNo',   None),
        (9,  'PROJECT NO.:',                 'PrjNumber',    v['contract_label'] + ' DATE:', 'ContractDate', DATE),
        (10, v['via_label'] + ':',           'ThirdName',    'ALLEATO PROJECT ID:',          'PrjId',        None),
    ]
    for r, lbl, name, rlbl, rname, fmt in party:
        c = ws.cell(r, 2, lbl)
        c.font = F_LBL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        vc = ws.cell(r, 4, f'=IF({name}="","",{name})')
        vc.font = F_BODY
        vc.border = Border(bottom=THIN)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        rc = ws.cell(r, 6, rlbl)
        rc.font = F_LBL
        rc.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        rv = ws.cell(r, 8, f'=IF({rname}="","",{rname})')
        rv.font = F_BODY
        rv.border = Border(bottom=THIN)
        if fmt:
            rv.number_format = fmt
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)

    banner(ws, 12, 2, 5, v['payee'] + '’S APPLICATION FOR PAYMENT')
    ws['B13'] = 'Application is made for payment, as shown below, in connection with the Contract.'
    ws['B13'].font = F_SMALL
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=5)

    LINES = [
        (15, '1',  'ORIGINAL CONTRACT SUM',                                                    '=OrigContractSum',   'calc'),
        (16, '2',  'Net change by Change Orders',                                              '=CO_NetTotal',       'calc'),
        (17, '3',  'CONTRACT SUM TO DATE  (Line 1 ± Line 2)',                                  '=N(E15)+N(E16)',     'bold'),
        (18, '4',  'TOTAL COMPLETED AND STORED TO DATE  (Continuation Sheet, column G)',                      '=SOV_TotalCompleted', 'bold'),
        (19, '5',  'RETAINAGE:',                                                               None,                 'section'),
        (20, '5a', 'of Completed Work  (SOV prior + this period − released)',                 '=SOV_RetWorkPrior+SOV_RetWorkThis-SOV_RetWorkReleased', 'calc'),
        (21, '5b', 'of Stored Material  (SOV prior + this period − released)',                '=SOV_RetMatlPrior+SOV_RetMatlThis-SOV_RetMatlReleased', 'calc'),
        (22, '',   'Total Retainage  (Lines 5a + 5b = SOV column S)',                         '=SOV_RetTotal',      'bold'),
        (23, '6',  'TOTAL EARNED LESS RETAINAGE  (Line 4 less Line 5 Total)',                  '=N(E18)-N(E22)',     'bold'),
        (24, '7',  'LESS PREVIOUS CERTIFICATES FOR PAYMENT  (Line 6 from prior Certificate)',  0,                    'input'),
        (25, '8',  'CURRENT PAYMENT DUE',                                                      '=N(E23)-N(E24)',     'total'),
        (26, '9',  'BALANCE TO FINISH, INCLUDING RETAINAGE  (Line 3 less Line 6)',             '=N(E17)-N(E23)',     'bold'),
    ]
    for r, no, label, formula, kind in LINES:
        n = ws.cell(r, 2, no)
        n.font = F_LBL
        n.alignment = Alignment(horizontal='center')
        lc = ws.cell(r, 3, label)
        if kind == 'section':
            lc.font = F_TOT
            for col in (2, 3, 4, 5):
                ws.cell(r, col).fill = SUB_FILL
            continue
        lc.font = F_TOT if kind in ('bold', 'total') else F_BODY
        ec = ws.cell(r, 5, formula)
        ec.number_format = MONEY
        ec.border = BOX
        if kind == 'input':
            ec.fill = IN_FILL
            ec.font = F_BODY
        elif kind == 'total':
            ec.fill = TOT_FILL
            ec.font = F_BIG
            lc.font = F_BIG
            ec.border = Border(top=MED, bottom=MED, left=MED, right=MED)
        elif kind == 'bold':
            ec.fill = LOCK_FILL
            ec.font = F_TOT
        else:
            ec.fill = LOCK_FILL
            ec.font = F_BODY

    for r, ref in ((20, '=RetPctWork'), (21, '=RetPctMatl')):
        p = ws.cell(r, 4, ref)
        p.number_format = PCT100
        p.font = F_TOT
        p.alignment = Alignment(horizontal='right')
        ws.cell(r, 3).alignment = Alignment(indent=2)

    rt = ws.cell(22, 7, '=IF(ROUND(N(E20)+N(E21)-N(E22),2)=0,'
                        '"Retainage 5a + 5b ties to SOV column S.",'
                        '"WARNING: retainage does not foot to SOV column S.")')
    rt.font = Font(name='Calibri', size=9, bold=True)
    rt.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=22, start_column=7, end_row=22, end_column=9)

    wb.defined_names['APP_ContractSumToDate'] = DefinedName(
        'APP_ContractSumToDate', attr_text=f"'{SH_APP}'!$E$17")
    wb.defined_names['APP_TotalEarnedLessRet'] = DefinedName(
        'APP_TotalEarnedLessRet', attr_text=f"'{SH_APP}'!$E$23")
    wb.defined_names['APP_CurrentPaymentDue'] = DefinedName(
        'APP_CurrentPaymentDue', attr_text=f"'{SH_APP}'!$E$25")

    # ---- change order summary
    banner(ws, 12, 7, 9, 'CHANGE ORDER SUMMARY')
    ws['H14'] = 'ADDITIONS'
    ws['I14'] = 'DEDUCTIONS'
    for c in ('H14', 'I14'):
        ws[c].font = F_HDR
        ws[c].fill = HDR_FILL
        ws[c].alignment = Alignment(horizontal='center')
        ws[c].border = BOX
    for r, lbl, a, d in [
        (15, 'Total changes approved in previous applications by ' + v['payer'].title(),
         '=CO_PrevAdd', '=CO_PrevDed'),
        (16, 'Total approved this application', '=CO_ThisAdd', '=CO_ThisDed'),
    ]:
        c = ws.cell(r, 7, lbl)
        c.font = F_BODY
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[r].height = 28
        for col, f in ((8, a), (9, d)):
            style_calc(ws.cell(r, col, f))

    ws.cell(17, 7, 'TOTALS').font = F_TOT
    for col in (8, 9):
        L = get_column_letter(col)
        t = ws.cell(17, col, f'=SUM({L}15:{L}16)')
        t.number_format = MONEY
        t.font = F_TOT
        t.fill = SUB_FILL
        t.border = BOX
    ws.cell(18, 7, 'NET CHANGE BY CHANGE ORDERS').font = F_TOT
    net = ws.cell(18, 8, '=N(H17)-N(I17)')
    net.number_format = MONEY
    net.font = F_TOT
    net.fill = TOT_FILL
    net.border = Border(top=MED, bottom=MED, left=MED, right=MED)
    ws.merge_cells(start_row=18, start_column=8, end_row=18, end_column=9)

    tie = ws.cell(20, 7, '=IF(ROUND(N(H18)-N(E16),2)=0,'
                         '"Net change ties to Line 2.",'
                         '"WARNING: Line 2 does not tie to the change order summary.")')
    tie.font = Font(name='Calibri', size=9, bold=True)
    tie.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=20, start_column=7, end_row=21, end_column=9)

    ws.cell(23, 7, 'PERCENT COMPLETE').font = F_LBL
    pc = ws.cell(23, 9, '=IF(N(E17)=0,"",N(E18)/N(E17)*100)')
    pc.number_format = PCT100
    pc.font = F_TOT
    pc.fill = LOCK_FILL
    pc.border = BOX
    ws.cell(24, 7, 'RETAINAGE HELD TO DATE').font = F_LBL
    rh = ws.cell(24, 9, '=N(E22)')
    rh.number_format = MONEY
    rh.font = F_TOT
    rh.fill = LOCK_FILL
    rh.border = BOX

    # ---- payee certification
    r = 29
    banner(ws, r, 2, 9, v['payee'] + '’S CERTIFICATION')
    cert = (f'The undersigned {v["payee"].lower()} certifies that, to the best of its knowledge, information '
            f'and belief, the Work covered by this Application for Payment has been completed in accordance '
            f'with the Contract Documents, that all amounts have been paid by the {v["payee"].lower()} for '
            f'Work for which previous Certificates for Payment were issued and payments received from the '
            f'{v["payer"].lower()}, and that the current payment shown herein is now due.')
    c = ws.cell(r + 1, 2, cert)
    c.font = F_BODY
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 3, end_column=9)
    for rr in (r + 1, r + 2, r + 3):
        ws.row_dimensions[rr].height = 15

    sig = r + 5
    for i, lbl in enumerate([v['payee'] + ' (firm):', 'By (signature):', 'Printed name / title:', 'Date:']):
        rr = sig + i * 2
        lc = ws.cell(rr, 2, lbl)
        lc.font = F_LBL
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
        f = ws.cell(rr, 4)
        f.border = Border(bottom=THIN)
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=5)

    ws.cell(sig, 7, 'State of:').font = F_LBL
    ws.cell(sig + 2, 7, 'County of:').font = F_LBL
    nt = ws.cell(sig + 4, 7, 'Subscribed and sworn to before me this ____ day of ____________, 20____.')
    nt.font = F_SMALL
    nt.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=sig + 4, start_column=7, end_row=sig + 4, end_column=9)
    ws.cell(sig + 6, 7, 'Notary Public:').font = F_LBL
    for rr in (sig, sig + 2, sig + 6):
        ws.cell(rr, 9).border = Border(bottom=THIN)

    # ---- certifier block
    r2 = sig + 9
    banner(ws, r2, 2, 9, v['cert_title'])
    cert2 = (f'In accordance with the Contract Documents and based on observations at the site and the data '
             f'comprising this Application, the {v["certifier"].lower()} certifies to the {v["payer"].lower()} '
             f'that the Work has progressed as indicated, that the quality of the Work is in accordance with '
             f'the Contract Documents, and that the {v["payee"].lower()} is entitled to payment of the '
             f'AMOUNT CERTIFIED.')
    c = ws.cell(r2 + 1, 2, cert2)
    c.font = F_BODY
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r2 + 1, start_column=2, end_row=r2 + 3, end_column=9)

    ac = ws.cell(r2 + 5, 2, 'AMOUNT CERTIFIED')
    ac.font = F_BIG
    ws.merge_cells(start_row=r2 + 5, start_column=2, end_row=r2 + 5, end_column=3)
    amt = ws.cell(r2 + 5, 5, '=APP_CurrentPaymentDue')
    amt.number_format = MONEY
    amt.font = F_BIG
    amt.fill = TOT_FILL
    amt.border = Border(top=MED, bottom=MED, left=MED, right=MED)

    ex = ws.cell(r2 + 6, 2, 'If AMOUNT CERTIFIED differs from CURRENT PAYMENT DUE, attach an explanation and '
                            'initial every figure changed on this Application and on the Continuation Sheet.')
    ex.font = F_SMALL
    ex.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r2 + 6, start_column=2, end_row=r2 + 6, end_column=9)
    for col in range(2, 10):
        ws.cell(r2 + 7, col).border = Border(bottom=THIN)

    for i, lbl in enumerate([v['certifier'] + ' (firm):', 'By (signature):', 'Date:']):
        rr = r2 + 9 + i * 2
        lc = ws.cell(rr, 2, lbl)
        lc.font = F_LBL
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
        f = ws.cell(rr, 4)
        f.border = Border(bottom=THIN)
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=5)

    fin = ws.cell(r2 + 16, 2,
                  'This Certificate is not negotiable. The AMOUNT CERTIFIED is payable only to the '
                  + v['payee'].lower() + ' named herein. Issuance, payment and acceptance of payment are '
                  'without prejudice to any rights of the ' + v['payer'].lower() + ' or '
                  + v['payee'].lower() + ' under the Contract.')
    fin.font = F_SMALL
    fin.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r2 + 16, start_column=2, end_row=r2 + 17, end_column=9)

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ------------------------------------------------------------ 5. Change Order
def build_change_order(wb, v, variant):
    ws = wb.create_sheet(SH_CO)
    ws.sheet_view.showGridLines = False
    for k, w in {'A': 3, 'B': 30, 'C': 26, 'D': 14, 'E': 12, 'F': 18, 'G': 18}.items():
        ws.column_dimensions[k].width = w

    brand_mast(ws, 7, accent_row=3, logo_col=5, logo_h=40)
    brand_print(ws, 'CO')
    ws['B1'] = 'CHANGE ORDER'
    ws['B1'].font = F_TITLE
    ws['B2'] = 'Form CO  ·  ' + v['co_title']
    ws['B2'].font = F_SUB

    for i, (lbl, val, rlbl, rval, fmt) in enumerate([
        ('PROJECT:', '=IF(PrjName="","",PrjName)', 'CHANGE ORDER NO.:', None, None),
        ('PROJECT NO.:', '=IF(PrjNumber="","",PrjNumber)', 'DATE:', None, DATE),
        (v['payer'] + ':', '=IF(PayerName="","",PayerName)', v['contract_label'] + ' NO.:', '=IF(ContractNo="","",ContractNo)', None),
        (v['payee'] + ':', '=IF(PayeeName="","",PayeeName)', v['contract_label'] + ' DATE:', '=IF(ContractDate="","",ContractDate)', DATE),
        (v['third_role'].upper() + ':', '=IF(ThirdName="","",ThirdName)', 'ALLEATO PROJECT ID:', '=IF(PrjId="","",PrjId)', None),
    ]):
        r = 4 + i
        ws.cell(r, 2, lbl).font = F_LBL
        c = ws.cell(r, 3, val)
        c.font = F_BODY
        c.border = Border(bottom=THIN)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        rc = ws.cell(r, 5, rlbl)
        rc.font = F_LBL
        rc.alignment = Alignment(horizontal='right')
        rv = ws.cell(r, 6, rval)
        rv.font = F_BODY
        rv.border = Border(bottom=THIN)
        if fmt:
            rv.number_format = fmt
        if rval is None:
            rv.fill = IN_FILL
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    banner(ws, 10, 2, 7, 'THE CONTRACT IS CHANGED AS FOLLOWS')
    ws['B11'] = ('Describe the change in scope, the reason for the change, and reference the authorizing '
                 'document (RFI, ASI, bulletin, field directive, CCD).')
    ws['B11'].font = F_SMALL
    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=7)
    for r in range(12, 20):
        for col in range(2, 8):
            ws.cell(r, col).border = Border(bottom=THIN)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

    banner(ws, 21, 2, 7, 'COST BREAKDOWN')
    for i, h in enumerate(['DESCRIPTION', 'BUDGET CODE', 'QTY', 'UOM', 'UNIT COST', 'AMOUNT']):
        c = ws.cell(22, 2 + i, h)
        c.font = F_HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')
        c.border = BOX

    L_FIRST, L_LAST = 23, 34
    for r in range(L_FIRST, L_LAST + 1):
        style_input(ws.cell(r, 2))
        style_input(ws.cell(r, 3))
        style_input(ws.cell(r, 4), '#,##0.00')
        style_input(ws.cell(r, 5))
        style_input(ws.cell(r, 6), MONEY)
        style_calc(ws.cell(r, 7, f'=IF(COUNT(D{r},F{r})=2,N(D{r})*N(F{r}),"")'))

    tr = L_LAST + 1
    ws.cell(tr, 2, 'TOTAL THIS CHANGE ORDER').font = F_TOT
    for col in range(2, 8):
        ws.cell(tr, col).fill = TOT_FILL
        ws.cell(tr, col).border = Border(top=MED, bottom=MED, left=THIN, right=THIN)
    t = ws.cell(tr, 7, f'=SUM(G{L_FIRST}:G{L_LAST})')
    t.number_format = MONEY
    t.font = F_TOT
    ws.cell(tr + 1, 2, 'If AMOUNT is entered directly without qty × unit cost, put it in the AMOUNT column '
                       'and leave QTY / UNIT COST blank.').font = F_SMALL
    ws.merge_cells(start_row=tr + 1, start_column=2, end_row=tr + 1, end_column=7)

    r = tr + 3
    banner(ws, r, 2, 7, 'CONTRACT SUM ADJUSTMENT')
    recon = [
        ('The original Contract Sum was',                                                    '=OrigContractSum', 'calc'),
        ('Net change by previously authorized Change Orders',                                0,                  'input'),
        ('The Contract Sum prior to this Change Order was',                                  f'=N(F{r+1})+N(F{r+2})', 'bold'),
        ('The Contract Sum will be increased / (decreased) by this Change Order in the amount of', f'=N(G{tr})',  'calc'),
        ('The new Contract Sum including this Change Order will be',                          f'=N(F{r+3})+N(F{r+4})', 'total'),
    ]
    for i, (lbl, f, kind) in enumerate(recon):
        rr = r + 1 + i
        c = ws.cell(rr, 2, lbl)
        c.font = F_TOT if kind in ('bold', 'total') else F_BODY
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
        ws.row_dimensions[rr].height = 22
        vc = ws.cell(rr, 6, f)
        vc.number_format = MONEY
        vc.border = BOX
        ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=7)
        if kind == 'input':
            vc.fill = IN_FILL
            vc.font = F_BODY
        elif kind == 'total':
            vc.fill = TOT_FILL
            vc.font = F_BIG
            vc.border = Border(top=MED, bottom=MED, left=MED, right=MED)
        elif kind == 'bold':
            vc.fill = LOCK_FILL
            vc.font = F_TOT
        else:
            vc.fill = LOCK_FILL
            vc.font = F_BODY

    r2 = r + 7
    for i, (lbl, suffix, fmt) in enumerate([
        ('The Contract Time will be increased / (decreased) by', 'calendar days', '#,##0'),
        ('The date of Substantial Completion as of this Change Order is', '', DATE),
    ]):
        rr = r2 + i
        c = ws.cell(rr, 2, lbl)
        c.font = F_BODY
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
        vc = ws.cell(rr, 6)
        vc.border = BOX
        vc.fill = IN_FILL
        vc.number_format = fmt
        ws.cell(rr, 7, suffix).font = F_SMALL

    note = ws.cell(r2 + 3, 2,
                   'NOTE: This Change Order does not include changes in the Contract Sum, Contract Time or '
                   'Guaranteed Maximum Price that have been authorized by Construction Change Directive '
                   'until the cost and time have been agreed upon by both parties.')
    note.font = F_SMALL
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r2 + 3, start_column=2, end_row=r2 + 4, end_column=7)

    sr = r2 + 6
    ws.cell(sr, 2, 'NOT VALID UNTIL SIGNED BY ALL PARTIES BELOW').font = F_TOT
    ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=7)

    parties = v['co_parties']
    span = 2 if len(parties) == 3 else 3
    for idx, role in enumerate(parties):
        c0 = 2 + idx * span
        ws.cell(sr + 2, c0, role).font = F_LBL
        for j, lbl in enumerate(['(firm name)', 'BY  (signature)', '(printed name and title)', 'DATE']):
            rr = sr + 3 + j * 3
            f = ws.cell(rr, c0)
            f.border = Border(bottom=THIN)
            if span > 1:
                ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c0 + span - 1)
            lab = ws.cell(rr + 1, c0, lbl)
            lab.font = F_SMALL

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ------------------------------------------------------------------- 6. README
def build_readme(wb, v, variant):
    ws = wb.create_sheet('README', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 114

    brand_mast(ws, 2, accent_row=1, logo_col=2, logo_h=46)
    brand_print(ws)
    ws['B2'] = 'ALLEATO GROUP BILLING TEMPLATE — ' + (
        'OWNER BILLING' if variant == 'owner' else 'SUBCONTRACTOR BILLING')
    ws['B2'].font = F_TITLE
    ws['B3'] = BRAND_TAGLINE + '     ·     ' + BRAND_WEB
    ws['B3'].font = Font(name='Calibri', size=10, italic=True, color=ORANGE2)

    body = [
        ('h', 'WHAT THIS IS'),
        ('p', 'Alleato Group\u2019s standard billing package: Form APP (Application and Certificate '
              'for Payment), Form CS (Continuation Sheet) and Form CO (Change Order), built as live '
              'formulas so a pay application can be produced, checked and archived without retyping '
              'numbers.'),
        ('p', 'Direction of billing: ' + v['payee'].title() + ' \u2192 ' + v['payer'].title() + '.'),
        ('h', 'SHEETS, IN THE ORDER YOU USE THEM'),
        ('li', 'Setup \u2014 project, parties, contract sum, retainage rates, current application '
               'period. The only sheet where project data is typed.'),
        ('li', 'CO Log \u2014 every change order. Status must read Approved or Executed to be picked '
               'up. The In Prior Application Y/N flag splits the change order summary between prior '
               'applications and this one.'),
        ('li', 'Schedule of Values \u2014 the working sheet and the monthly data entry: Scheduled '
               'Value (col D), From Previous Application (col E), This Period (col F), Materials '
               'Presently Stored (col G), and the prior / released retainage columns. Everything '
               'else calculates. Wider than the Continuation Sheet because it carries the retainage '
               'detail the platform needs.'),
        ('li', 'Continuation Sheet \u2014 Form CS. The same schedule of values in the standard '
               'columns A through I. Every cell is a formula from the Schedule of Values sheet; '
               'nothing is typed here. This is the copy you send out.'),
        ('li', 'Payment Application \u2014 Form APP. Reads entirely from the Schedule of Values and '
               'CO Log. The only typed cell is Line 7, Less Previous Certificates for Payment.'),
        ('li', 'Change Order \u2014 Form CO. One change order per sheet; duplicate the sheet as '
               'needed.'),
        ('h', 'RETAINAGE \u2014 5a / 5b SPLIT'),
        ('p', 'Retainage is tracked per SOV line in three parts, matching how the platform stores '
              'it: PRIOR (withheld on earlier applications), THIS PERIOD (calculated from the rate), '
              'and RELEASED. Completed work and stored materials are tracked separately, which is '
              'what produces Line 5a and Line 5b.'),
        ('p', 'Columns K and O hold the rates and default from Setup \u2014 override either on an '
              'individual line when that line carries a reduced rate. Columns M and Q calculate this '
              'period\u2019s retainage. Columns L and P are typed: carry them forward from the prior '
              'application\u2019s held balance.'),
        ('p', 'To release retainage, enter the released amount as a POSITIVE number in column N '
              '(work) or column R (materials). Never adjust Line 5 directly \u2014 it must always '
              'foot to column S, and the sheet warns you if it does not.'),
        ('p', 'All percentages in this workbook are entered out of 100 (type 10 for ten percent), '
              'matching the platform\u2019s retainage_pct / work_completed_pct columns. Do not '
              'enter 0.10.'),
        ('h', 'MONTHLY ROLLOVER'),
        ('li', 'Save a copy of the workbook named for the new application number.'),
        ('li', 'On Setup, bump Application No. and set the new Period From / Period To.'),
        ('li', 'On the Schedule of Values, copy column H (Total Completed and Stored to Date) and '
               'paste-special as VALUES into column E (From Previous Application). Then clear '
               'columns F and G.'),
        ('li', 'Also on the Schedule of Values, paste last month\u2019s held retainage into columns '
               'L and P as VALUES (work: L + M \u2212 N; materials: P + Q \u2212 R), then clear '
               'columns N and R.'),
        ('li', 'On the Payment Application, set Line 7 to the prior application\u2019s Line 6.'),
        ('li', 'On the CO Log, flip In Prior Application to Y for change orders that were on the '
               'last application.'),
        ('h', 'BUILT-IN CHECKS'),
        ('li', 'Tie-out: the Scheduled Value grand total must equal Line 3, Contract Sum to Date. '
               'The cell below the grand total reports OK or the variance.'),
        ('li', 'Change order summary: net change must tie to Line 2. A warning appears if it does '
               'not.'),
        ('li', 'Retainage: Lines 5a + 5b must foot to column S. A warning appears if they do not.'),
        ('h', 'UPLOADING TO THE ALLEATO PM PLATFORM'),
        ('p', 'The CSV templates shipped alongside this workbook map these sheets to the platform '
              'tables. The Schedule of Values columns line up one-for-one with the invoice line item '
              'columns, so a completed sheet can be copied straight into the CSV. See '
              'CSV-IMPORT-SPEC.md for the column mapping and required load order.'),
        ('h', 'SCOPE OF THESE FORMS'),
        ('p', 'These are Alleato Group forms. They follow the standard construction-industry '
              'application-for-payment structure \u2014 nine summary lines, the 5a / 5b retainage '
              'split, and a continuation sheet with columns A through I \u2014 so any owner, lender '
              'or subcontractor will read them without friction. Certification language is Alleato\u2019s '
              'own. Where a contract specifies submission on a particular third-party form, obtain '
              'that form from its publisher and key the figures across from this workbook; use these '
              'forms for internal billing preparation, for subcontractor billing, and as the data '
              'source for the platform.'),
    ]
    r = 4
    for kind, text in body:
        c = ws.cell(r, 2, ('•  ' + text) if kind == 'li' else text)
        if kind == 'h':
            c.font = F_SECT
            c.fill = HDR_FILL
            r += 1
            continue
        c.font = F_BODY
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 15 * max(1, (len(text) // 105) + 1)
        r += 1
    r += 1
    for col in (1, 2):
        ws.cell(r, col).fill = ACCENT_FILL
    ws.row_dimensions[r].height = 3.5
    ws.cell(r + 1, 2, 'ALLEATO GROUP  ·  ' + BRAND_WEB).font = Font(
        name='Calibri', size=10, bold=True, color=DKGREY)
    ws.cell(r + 2, 2, BRAND_HQ).font = F_SMALL
    ws.cell(r + 3, 2, BRAND_FL).font = F_SMALL
    ws.cell(r + 4, 2, BRAND_TAGLINE).font = Font(
        name='Calibri', size=10, italic=True, color=ORANGE2)


def build(variant, path):
    v = variant_words(variant)
    wb = Workbook()
    wb.remove(wb.active)
    build_setup(wb, v, variant)
    build_co_log(wb, v)
    build_schedule_of_values(wb, v, variant)
    build_continuation_sheet(wb, v, variant)
    build_payment_application(wb, v, variant)
    build_change_order(wb, v, variant)
    build_readme(wb, v, variant)
    wb.active = 0
    wb.save(path)
    print('wrote', path)


if __name__ == '__main__':
    out = sys.argv[1]
    build('owner', f'{out}/Alleato-Billing-Template-OWNER.xlsx')
    build('sub', f'{out}/Alleato-Billing-Template-SUBCONTRACTOR.xlsx')
