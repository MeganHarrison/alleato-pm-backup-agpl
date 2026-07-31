#!/usr/bin/env python3
"""Populate the OWNER workbook with a worked example so LibreOffice can recalc it."""
import re
import sys
from openpyxl import load_workbook

src, dst = sys.argv[1], sys.argv[2]
wb = load_workbook(src)


def ref(name):
    dn = wb.defined_names[name]
    m = re.match(r"^'?([^'!]+)'?!\$([A-Z]+)\$(\d+)$", dn.attr_text)
    sheet, col, row = m.group(1), m.group(2), int(m.group(3))
    return wb[sheet], f'{col}{row}'


def put(name, value):
    ws, addr = ref(name)
    ws[addr] = value


put('PrjName', 'Morrisville Distribution Center')
put('PrjNumber', '2026-118')
put('PrjId', 876)
put('PayerName', 'Northpoint Development LLC')
put('PayeeName', 'Alleato Group, LLC')
put('ThirdName', 'Harmon Architects')
put('ContractNo', 'PC-2026-118-001')
put('OrigContractSum', 1_000_000)
put('RetPctWork', 10.0)
put('RetPctMatl', 10.0)
put('AppNo', 4)

co = wb['CO Log']
for r, (num, desc, status, add, ded, days, prior) in enumerate([
    ('CO-001', 'Added dock levelers', 'Approved', 50_000, 0, 0, 'Y'),
    ('CO-002', 'Revised roof insulation / deleted skylights', 'Approved', 30_000, 5_000, 5, 'N'),
], start=7):
    co[f'A{r}'], co[f'B{r}'], co[f'D{r}'] = num, desc, status
    co[f'E{r}'], co[f'F{r}'], co[f'G{r}'], co[f'H{r}'] = add, ded, days, prior

g = wb['Schedule of Values']
base = [
    ('1', 'General Conditions', '01-1000', 100_000, 40_000, 10_000, 0),
    ('2', 'Cast-in-Place Concrete', '03-3000', 400_000, 200_000, 50_000, 20_000),
    ('3', 'Structural Steel', '05-1200', 500_000, 0, 100_000, 30_000),
]
for i, (no, desc, code, sv, prev, this, stored) in enumerate(base):
    r = 12 + i
    g[f'A{r}'], g[f'B{r}'], g[f'C{r}'] = no, desc, code
    g[f'D{r}'], g[f'E{r}'], g[f'F{r}'], g[f'G{r}'] = sv, prev, this, stored
    g[f'L{r}'] = round(prev * 0.10, 2)     # retainage held on prior work

cos = [
    ('CO-001', 'Dock levelers', '11-1300', 50_000, 20_000, 10_000, 0),
    ('CO-002', 'Roof insulation revision', '07-2100', 25_000, 0, 5_000, 0),
]
for i, (no, desc, code, sv, prev, this, stored) in enumerate(cos):
    r = 50 + i
    g[f'A{r}'], g[f'B{r}'], g[f'C{r}'] = no, desc, code
    g[f'D{r}'], g[f'E{r}'], g[f'F{r}'], g[f'G{r}'] = sv, prev, this, stored
    g[f'L{r}'] = round(prev * 0.10, 2)

wb['Payment Application']['E24'] = 240_000   # Line 7 — less previous certificates

wb.save(dst)
print('populated ->', dst)
